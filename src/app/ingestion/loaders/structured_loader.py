from __future__ import annotations

import asyncio
import csv
import json
from io import BytesIO, StringIO
from pathlib import Path
import xml.etree.ElementTree as et

from bs4 import BeautifulSoup
from openpyxl import load_workbook

from app.contracts.ingestion import BaseLoader, DocumentBlob, ParsedDocument


def _flatten_xml(content: bytes) -> str:
    root = et.fromstring(content)
    values = []
    for element in root.iter():
        text = (element.text or "").strip()
        if text:
            values.append(f"{element.tag}: {text}")
    return "\n".join(values)


def _flatten_xlsx(content: bytes) -> str:
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(f"Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                lines.append(" | ".join(values))
    return "\n".join(lines)


class StructuredLoader(BaseLoader):
    name = "structured"
    _extensions = {".csv", ".json", ".xml", ".html", ".htm", ".xlsx"}

    def supports(self, filename: str, content_type: str) -> bool:
        ext = Path(filename).suffix.lower()
        if ext in self._extensions:
            return True
        return content_type in {
            "application/json",
            "application/xml",
            "text/csv",
            "text/html",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }

    async def load(self, blob: DocumentBlob) -> ParsedDocument:
        ext = Path(blob.filename).suffix.lower()
        text = ""

        if ext == ".csv":
            decoded = blob.content_bytes.decode("utf-8", errors="ignore")
            reader = csv.reader(StringIO(decoded))
            text = "\n".join(", ".join(col.strip() for col in row if col is not None) for row in reader)
        elif ext == ".json":
            payload = json.loads(blob.content_bytes.decode("utf-8", errors="ignore"))
            text = json.dumps(payload, indent=2, ensure_ascii=True)
        elif ext == ".xml":
            text = _flatten_xml(blob.content_bytes)
        elif ext in {".html", ".htm"}:
            soup = BeautifulSoup(blob.content_bytes, "lxml")
            text = soup.get_text(separator="\n", strip=True)
        elif ext == ".xlsx":
            text = await asyncio.to_thread(_flatten_xlsx, blob.content_bytes)
        else:
            text = blob.content_bytes.decode("utf-8", errors="ignore")

        return ParsedDocument(
            document_id=blob.document_id,
            text=text,
            metadata={
                **blob.metadata,
                "filename": blob.filename,
                "loader": self.name,
            },
        )
